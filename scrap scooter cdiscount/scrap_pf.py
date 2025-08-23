# -*- coding: utf-8 -*-
"""
Created on Sun Aug  3 16:42:19 2025

@author: Charles
"""

import random
import os
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import undetected_chromedriver as UC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.5845.111 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.5735.198 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/117.0.5938.62 Safari/537.36"
]

ua = random.choice(user_agents)

options = UC.ChromeOptions()
options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36 OPR/120.0.0.0 (Edition std-1)")


service = Service(executable_path="chromedriver-win32/chromedriver.exe")
#options = Options()

options.add_argument("--start-maximized")
#options.add_experimental_option("excludeSwitches", ["enable-automation"])
#options.add_experimental_option("useAutomationExtension", False)
options.add_argument("--lang=fr-FR")
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--headless")
os.environ["PATH"] += r"C:/SeleniumDrivers"

url = "https://www.cdiscount.com/search/10/trottinette+electrique.html#_his_"
driver = UC.Chrome(options=options)
driver.get(url)

html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

ct = []

attente = WebDriverWait(driver, 3)

#troti = []
#troti.append(driver.find_element(By.CSS_SELECTOR, '[data-e2e="search-result-block"]'))

#titres = driver.find_elements(By.CSS_SELECTOR, 'h4.o-card__title.u-line-clamp--2')

# En-têtes

ws.append(['Titre', 'Prix (€)', 'Note (/ 5)', 'Lien'])
somme_p = 0
somme_not = 0

compteur_p = 0
compteur_n = 0

for titre, prix, note, lien in zip(soup.select('[data-e2e="lplr-title"]'), soup.select('.sc-1vyswr-1.hPhsSW.price'), soup.select('.sc-dvQaRk.fGyZww'), soup.select('a[href].sc-1aa1prr-20.eyiUQR')):
    t = titre.get_text(strip=True) if titre else "N/A"
    p = prix.get_text(strip=True) if prix else "N/A"
    n = note.get_text(strip=True) if note else "N/A"
    l = lien['href'] if lien else "N/A"
    
    somme_p += float(p.replace('€', '').replace(',', '.'))
    somme_not += float(n.replace(',', '.').replace('/ 5', ''))
    compteur_p += 1
    compteur_n += 1
    
    ws.append([
        t,
        float(p.replace('€', '').replace(',', '.')),
        float(n.replace(',', '.').replace('/ 5', '')),
        l
    ])

moyenne_p = somme_p/compteur_p if compteur_p > 0 else 0
moyenne_n = somme_not/compteur_n if compteur_n > 0 else 0

ws["B1"].value = f"Prix (moyenne : {moyenne_p:.2f} €)"
ws["C1"].value = f"Note (moyenne : {moyenne_n:.1f} / 5)"

for cell in ws[1]:
    cell.font = Font(bold=True)

for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
    for cell in row:
        if cell.column == 2:
            if cell.value > moyenne_p:
                cell.fill = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid') # rouge
            cell.value = str(cell.value) + "€"
        elif cell.column == 3:
            if cell.value < moyenne_n:
                cell.fill = PatternFill(start_color='FF7F7F', end_color='FF7F7F', fill_type='solid') # rouge
            cell.value = str(cell.value) + "/ 5"

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        except:
            pass
    
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save("produits_cdiscount_colo.xlsx")

driver.quit()